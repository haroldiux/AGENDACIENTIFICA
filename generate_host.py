import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# Sheet 1: Actividades Académicas
ws_ac = wb.active
ws_ac.title = 'Actividades Académicas'
academic_headers = ['titulo', 'fecha_inicio', 'fecha_fin', 'categoria', 'carrera', 'gestion', 'es_cientifica']

fill_green = PatternFill(start_color='009E96', end_color='009E96', fill_type='solid')
font_head = Font(bold=True, color='FFFFFF', size=11)
border = Border(bottom=Side(border_style='thin', color='CCCCCC'), right=Side(border_style='thin', color='CCCCCC'))

for col_idx, h in enumerate(academic_headers, start=1):
    c = ws_ac.cell(row=1, column=col_idx, value=h)
    c.fill = fill_green
    c.font = font_head
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
    ws_ac.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

# Sheet 2: Actividades Científicas (10 rows)
ws_sc = wb.create_sheet(title='Actividades Científicas')
scientific_headers = ['titulo', 'fecha_inicio', 'fecha_fin', 'tipo_actividad', 'nombre_responsable', 'carrera', 'gestion', 'es_cientifica']
fill_purple = PatternFill(start_color='6B3392', end_color='6B3392', fill_type='solid')

for col_idx, h in enumerate(scientific_headers, start=1):
    c = ws_sc.cell(row=1, column=col_idx, value=h)
    c.fill = fill_purple
    c.font = font_head
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
    ws_sc.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

activities_data = [
    ('Congreso Internacional de Inteligencia Artificial y Robótica', '15/08/2026', '18/08/2026', 'CONGRESO', 'Dr. Roberto Carlos Flores', 'Ingeniería de Sistemas', '2-2026', 'SI'),
    ('Webinar: Innovaciones en Salud Digital y Telemedicina', '20/08/2026', '20/08/2026', 'WEBINAR', 'Dra. María Elena Gutiérrez', 'Medicina', '2-2026', 'SI'),
    ('Defensa de Tesis: Simulación e Impresión 3D Biomédica', '01/09/2026', '02/09/2026', 'DEFENSA', 'Ing. Marcelo Vargas', 'Ingeniería Biomédica', '2-2026', 'SI'),
    ('Feria de Ciencia y Tecnología UNITEPC 2026', '10/09/2026', '12/09/2026', 'FERIA', 'Lic. Sofía Mendoza', 'Ingeniería de Sonido', '2-2026', 'SI'),
    ('Olimpiada Nacional de Algoritmos y Programación', '18/09/2026', '20/09/2026', 'OLIMPIADA', 'MSc. Alejandro Paredes', 'Ingeniería de Sistemas', '2-2026', 'SI'),
    ('Master Class: Bioseguridad e Investigación Clínica', '05/10/2026', '06/10/2026', 'MASTER_CLASS', 'Dr. Carlos Fernando Suárez', 'Bioquímica y Farmacia', '2-2026', 'SI'),
    ('Simposio de Odontología Digital e Implantología Avanzada', '12/10/2026', '14/10/2026', 'CONGRESO', 'Dra. Patricia Morales', 'Odontología', '2-2026', 'SI'),
    ('Seminario de Bioética y Derecho en la Investigación', '22/10/2026', '23/10/2026', 'WEBINAR', 'Lic. Fernando Torrez', 'Derecho', '2-2026', 'SI'),
    ('Encuentro de Investigación Económica y Financiera', '05/11/2026', '07/11/2026', 'FERIA', 'Lic. Gonzalo Aguilar', 'Economía', '2-2026', 'SI'),
    ('Jornadas Científicas de Enfermería y Cuidados Críticos', '15/11/2026', '17/11/2026', 'CONGRESO', 'Dra. Claudia Ramos', 'Enfermería', '2-2026', 'SI'),
]

for row_idx, data in enumerate(activities_data, start=2):
    for col_idx, val in enumerate(data, start=1):
        cell = ws_sc.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal='center' if col_idx in (2,3,4,7,8) else 'left', vertical='center')
        if col_idx in (2, 3):
            cell.number_format = 'DD/MM/YYYY'

# Sheet 3: Referencia
ws_ref = wb.create_sheet(title='Referencia')
ws_ref.column_dimensions['E'].width = 35
ws_ref.column_dimensions['I'].width = 22
ws_ref.column_dimensions['L'].width = 25

ws_ref.cell(row=2, column=5, value='Nombre de la Carrera').font = Font(bold=True)
careers = ['Ingeniería de Sistemas', 'Medicina', 'Odontología', 'Derecho', 'Ingeniería Biomédica', 'Ingeniería de Sonido', 'Bioquímica y Farmacia', 'Economía', 'Enfermería']
for r_off, c in enumerate(careers, start=3):
    ws_ref.cell(row=r_off, column=5, value=c)

ws_ref.cell(row=2, column=9, value='Nombre de la Gestión').font = Font(bold=True)
gestiones = ['1-2024', '2-2024', '1-2025', '2-2025', '1-2026', '2-2026']
for r_off, g in enumerate(gestiones, start=3):
    ws_ref.cell(row=r_off, column=9, value=g)

ws_ref.cell(row=2, column=12, value='Nombre Categoría').font = Font(bold=True)
categories = ['CONGRESO', 'WEBINAR', 'DEFENSA', 'FERIA', 'OLIMPIADA', 'MASTER_CLASS']
for r_off, cat in enumerate(categories, start=3):
    ws_ref.cell(row=r_off, column=12, value=cat)

# Data Validations
dv_bool = DataValidation(type='list', formula1='"SI,NO"', allow_blank=True)
ws_sc.add_data_validation(dv_bool)
dv_bool.add('H2:H500')

dv_date = DataValidation(type='date', operator='greaterThanOrEqual', formula1='2000-01-01', allow_blank=True)
ws_sc.add_data_validation(dv_date)
dv_date.add('B2:C500')

dv_car = DataValidation(type='list', formula1=f"'Referencia'!E3:E{2 + len(careers)}", allow_blank=True)
ws_sc.add_data_validation(dv_car)
dv_car.add('F2:F500')

dv_ges = DataValidation(type='list', formula1=f"'Referencia'!I3:I{2 + len(gestiones)}", allow_blank=True)
ws_sc.add_data_validation(dv_ges)
dv_ges.add('G2:G500')

dv_cat = DataValidation(type='list', formula1=f"'Referencia'!L3:L{2 + len(categories)}", allow_blank=True)
ws_sc.add_data_validation(dv_cat)
dv_cat.add('D2:D500')

target_dir = r"F:\AGENDACIENTIFICA\EJEMPLO EXCEL"
os.makedirs(target_dir, exist_ok=True)
out_file = os.path.join(target_dir, "plantilla_actividades (1).xlsx")
wb.save(out_file)
print("SUCCESS_HOST: Saved directly to host path:", out_file)
