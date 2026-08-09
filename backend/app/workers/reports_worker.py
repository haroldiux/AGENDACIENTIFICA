import os
import uuid
from collections import defaultdict
from datetime import datetime

from app.core.celery_app import celery_app
from app.db.session import SessionLocal
from app.models.models import AcademicActivity, ScientificActivity, Career, Gestion
from app.schemas.schemas import ConflictItem
from app.services import conflict_service

from jinja2 import Environment, FileSystemLoader
import base64
import urllib.request


def _get_weasyprint_html():
    """Lazy import of WeasyPrint so the backend can start without GTK libs."""
    try:
        from weasyprint import HTML
        return HTML
    except OSError as exc:
        raise RuntimeError(
            "WeasyPrint dependencies are missing. Install GTK/Pango or run inside Docker."
        ) from exc

CAREER_IMAGES = {
    "Ingeniería de Sistemas": "https://picsum.photos/seed/sistemas/800/500",
    "Medicina": "https://picsum.photos/seed/medicina/800/500",
    "Odontología": "https://picsum.photos/seed/odontologia/800/500",
    "Administración de Empresas": "https://picsum.photos/seed/admin/800/500",
    "Derecho": "https://picsum.photos/seed/derecho/800/500",
    "Arquitectura": "https://picsum.photos/seed/arquitectura/800/500",
    "Comunicación Social": "https://picsum.photos/seed/comunicacion/800/500",
    "Ingeniería Comercial": "https://picsum.photos/seed/comercial/800/500",
}
DEFAULT_CAREER_IMAGE = "https://picsum.photos/seed/unitepc/800/500"

def fetch_image_as_base64(url: str) -> str:
    import httpx
    try:
        with httpx.Client(follow_redirects=True, timeout=15.0) as client:
            response = client.get(url, headers={'User-Agent': 'Mozilla/5.0'})
            response.raise_for_status()
            b64_data = base64.b64encode(response.content).decode('utf-8')
            return f"data:image/jpeg;base64,{b64_data}"
    except Exception as e:
        print(f"Error fetching image {url}: {e}")
        return ""


REPORTS_DIR = os.getenv("REPORTS_DIR", os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "reports")) if os.name == "nt" else "/app/reports")
try:
    os.makedirs(REPORTS_DIR, exist_ok=True)
except PermissionError:
    REPORTS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "reports"))
    os.makedirs(REPORTS_DIR, exist_ok=True)

# Templates directory in the container (relative to this file)
TEMPLATES_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "templates"))
jinja_env = Environment(loader=FileSystemLoader(TEMPLATES_DIR))

SPANISH_MONTHS = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

SPANISH_MONTH_ABBR = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
    7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic",
}

ACTIVITY_TYPE_LABELS = {
    "congreso": "Congreso",
    "webinar": "Webinar",
    "defensa": "Defensa",
    "feria": "Feria",
    "olimpiada": "Olimpiada",
    "master_class": "Master Class",
}

ACTIVITY_STATUS_LABELS = {
    "scheduled": "Programada",
    "in_progress": "En progreso",
    "completed": "Completada",
    "cancelled": "Cancelada",
}

MAX_NOTES_LENGTH = 500


def _month_label(year, month):
    return f"{SPANISH_MONTHS[month]} {year}"


def _format_short_date(value):
    return f"{value.day} {SPANISH_MONTH_ABBR[value.month]} {value.year}"


def _format_date_range(start_date, end_date):
    if start_date == end_date:
        return _format_short_date(start_date)
    return f"{_format_short_date(start_date)} – {_format_short_date(end_date)}"


def _clamp_notes(notes):
    text = notes or "—"
    if len(text) > MAX_NOTES_LENGTH:
        return text[:MAX_NOTES_LENGTH] + "…"
    return text


def _enum_value(value):
    return value.value if hasattr(value, "value") else value


def get_activity_label(val):
    v = _enum_value(val)
    return ACTIVITY_TYPE_LABELS.get(v, str(v))


def get_status_label(val):
    v = _enum_value(val)
    return ACTIVITY_STATUS_LABELS.get(v, str(v))


# We pass these helper functions into the templates via context
DEFAULT_CONTEXT = {
    "format_date_range": _format_date_range,
    "clamp_notes": _clamp_notes,
    "get_activity_label": get_activity_label,
    "get_status_label": get_status_label,
    "_month_label": _month_label,
}


def _conflict_month_group(conflicts):
    grouped = defaultdict(list)
    for item in conflicts:
        grouped[(item.scientific_start_date.year, item.scientific_start_date.month)].append(item)
    return grouped


def build_conflict_pdf(filepath, conflicts, career_id, gestion_id):
    grouped = _conflict_month_group(conflicts) if conflicts else {}
    grouped_list = []
    for (year, month) in sorted(grouped.keys()):
        grouped_list.append({
            "label": _month_label(year, month),
            "activities": grouped[(year, month)]
        })

    context = {
        **DEFAULT_CONTEXT,
        "grouped_conflicts": grouped_list,
        "current_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "subtitle": f"Carrera ID: {career_id} — Gestión ID: {gestion_id}"
    }

    template = jinja_env.get_template("conflict_report.html")
    html_out = template.render(**context)
    _get_weasyprint_html()(string=html_out).write_pdf(filepath)


def build_conflict_excel(conflicts, career_id, gestion_id):
    from openpyxl import Workbook

    filename = f"conflict_report_{uuid.uuid4().hex}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Conflictos"

    ws.append(["Reporte de Conflictos"])
    ws.append([f"Carrera ID: {career_id} — Gestión ID: {gestion_id}"])
    ws.append([f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append([
        "Título académico",
        "Fechas académicas",
        "Título científico",
        "Tipo científico",
        "Fechas científicas",
    ])

    for item in conflicts:
        scientific_type_label = get_activity_label(item.scientific_type)
        ws.append([
            item.academic_title,
            _format_date_range(item.academic_start_date, item.academic_end_date),
            item.scientific_title,
            scientific_type_label,
            _format_date_range(item.scientific_start_date, item.scientific_end_date),
        ])

    wb.save(filepath)
    return {"status": "completed", "file_path": filepath, "file_name": filename}


def build_research_agenda_pdf(filepath, activities, career_name, gestion_name, report_title="Calendario Académico y Científico"):
    # Determine the target year (default to current year if no activities)
    target_year = activities[0].start_date.year if activities else datetime.now().year

    grouped = defaultdict(list)
    if activities:
        for act in activities:
            grouped[(act.start_date.year, act.start_date.month)].append(act)

    career_image_url = CAREER_IMAGES.get(career_name, DEFAULT_CAREER_IMAGE)
    base64_image = fetch_image_as_base64(career_image_url)

    grouped_list = []
    for month in range(1, 13):
        year_month = (target_year, month)
        acts = grouped.get(year_month, [])
        # Only include months that have activities, or if you want all 12, keep it. 
        # The user's image shows ONLY the months that have activities basically (Jul-Dec), 
        # but 12 months is fine. We will sort them correctly.
        if acts or month in (1, 2, 3, 4, 5, 6): # Include standard semester months
            grouped_list.append({
                "month_name": SPANISH_MONTHS[month].upper(),
                "month_idx": month,
                "activities": sorted(acts, key=lambda x: x.start_date),
                "side": "left" if len(grouped_list) % 2 == 0 else "right" # Alternate based on visible rows
            })

    context = {
        **DEFAULT_CONTEXT,
        "grouped_agenda": grouped_list,
        "current_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "year": gestion_name,
        "report_title": report_title,
        "career_name": career_name,
        "career_image": base64_image,
        "gestion_name": gestion_name
    }

    template = jinja_env.get_template("research_agenda.html")
    html_out = template.render(**context)
    _get_weasyprint_html()(string=html_out).write_pdf(filepath)


def _build_table_report(filepath, career_id, gestion_id, db, status_filter=None):
    ac_query = db.query(AcademicActivity)
    sc_query = db.query(ScientificActivity)

    if career_id:
        ac_query = ac_query.filter(AcademicActivity.career_id == career_id)
        sc_query = sc_query.filter(ScientificActivity.career_id == career_id)

    if gestion_id:
        ac_query = ac_query.filter(AcademicActivity.gestion_id == gestion_id)
        sc_query = sc_query.filter(ScientificActivity.gestion_id == gestion_id)

    if status_filter:
        sc_query = sc_query.filter(ScientificActivity.status == status_filter)
        if status_filter != 'scheduled':
            academic_activities = []
        else:
            academic_activities = ac_query.all()
    else:
        academic_activities = ac_query.all()

    scientific_activities = sc_query.all()

    status_filter_label = None
    if status_filter:
        status_filter_label = get_status_label(status_filter)

    context = {
        **DEFAULT_CONTEXT,
        "academic_activities": academic_activities,
        "scientific_activities": scientific_activities,
        "status_filter_label": status_filter_label,
        "current_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "subtitle": f"Carrera ID: {career_id or 'Todas'} — Gestión ID: {gestion_id}"
    }

    template = jinja_env.get_template("activity_report.html")
    html_out = template.render(**context)
    _get_weasyprint_html()(string=html_out).write_pdf(filepath)


def build_seguimiento_data(db, career_id, gestion_id):
    careers_all = db.query(Career).all()
    gestion = db.query(Gestion).filter(Gestion.id == gestion_id).first() if gestion_id else None
    gestion_name = gestion.name if gestion else "Gestión actual"

    if career_id:
        c_list = [c for c in careers_all if c.id == career_id]
        career_name = c_list[0].name if c_list else f"Carrera {career_id}"
    else:
        c_list = careers_all
        career_name = "Todas las carreras"

    ac_query = db.query(AcademicActivity)
    sc_query = db.query(ScientificActivity)

    if career_id is not None:
        ac_query = ac_query.filter(AcademicActivity.career_id == career_id)
        sc_query = sc_query.filter(ScientificActivity.career_id == career_id)
    if gestion_id is not None:
        ac_query = ac_query.filter(AcademicActivity.gestion_id == gestion_id)
        sc_query = sc_query.filter(ScientificActivity.gestion_id == gestion_id)

    ac_items = ac_query.all()
    sc_items = sc_query.all()

    ac_by_career = defaultdict(list)
    for a in ac_items:
        ac_by_career[a.career_id].append(a)

    sc_by_career = defaultdict(list)
    for s in sc_items:
        sc_by_career[s.career_id].append(s)

    all_targets = [None] + [c.id for c in c_list] if career_id is None else [c.id for c in c_list]
    career_id_to_name = {c.id: c.name for c in careers_all}
    career_id_to_name[None] = "Global / Vicerrectorado"

    careers_summary = []
    tot_total = 0
    tot_completed = 0
    tot_in_progress = 0
    tot_cancelled = 0
    tot_scheduled = 0

    for cid in all_targets:
        c_name = career_id_to_name.get(cid, "Otra")
        c_ac = ac_by_career.get(cid, [])
        c_sc = sc_by_career.get(cid, [])

        tot = len(c_ac) + len(c_sc)
        if tot == 0 and career_id is None:
            continue

        comp = sum(1 for s in c_sc if _enum_value(s.status) == "completed")
        inp = sum(1 for s in c_sc if _enum_value(s.status) == "in_progress")
        canc = sum(1 for s in c_sc if _enum_value(s.status) == "cancelled")
        sched = len(c_ac) + sum(1 for s in c_sc if _enum_value(s.status) == "scheduled")

        rate = round((comp / tot * 100), 1) if tot > 0 else 0.0

        tot_total += tot
        tot_completed += comp
        tot_in_progress += inp
        tot_cancelled += canc
        tot_scheduled += sched

        careers_summary.append({
            "career_id": cid,
            "career_name": c_name,
            "total": tot,
            "scheduled": sched,
            "in_progress": inp,
            "completed": comp,
            "cancelled": canc,
            "completion_rate": rate,
        })

    tot_rate = round((tot_completed / tot_total * 100), 1) if tot_total > 0 else 0.0
    totals = {
        "total": tot_total,
        "scheduled": tot_scheduled,
        "in_progress": tot_in_progress,
        "completed": tot_completed,
        "cancelled": tot_cancelled,
        "completion_rate": tot_rate,
    }

    detailed = []
    for s in sc_items:
        cname = career_id_to_name.get(s.career_id, "Global / Vicerrectorado")
        st_val = _enum_value(s.status)
        detailed.append({
            "title": s.title,
            "career_name": cname,
            "dates": _format_date_range(s.start_date, s.end_date),
            "type_label": get_activity_label(s.activity_type),
            "status": st_val,
            "status_label": get_status_label(s.status),
            "notes": s.notes,
            "has_evidence": len(s.evidences) > 0 or bool(s.evidence_url),
        })

    for a in ac_items:
        cname = career_id_to_name.get(a.career_id, "Global / Vicerrectorado")
        detailed.append({
            "title": a.title,
            "career_name": cname,
            "dates": _format_date_range(a.start_date, a.end_date),
            "type_label": a.category or "Académica",
            "status": "scheduled",
            "status_label": "Programada",
            "notes": None,
            "has_evidence": False,
        })

    return {
        "gestion_name": gestion_name,
        "career_name": career_name,
        "now_str": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "careers_summary": careers_summary,
        "totals": totals,
        "detailed_activities": detailed,
    }


def build_seguimiento_pdf(filepath, career_id, gestion_id, db):
    data = build_seguimiento_data(db, career_id, gestion_id)
    context = {
        **DEFAULT_CONTEXT,
        **data,
    }
    template = jinja_env.get_template("seguimiento_report.html")
    html_out = template.render(**context)
    _get_weasyprint_html()(string=html_out).write_pdf(filepath)


@celery_app.task
def generate_pdf_report_task(
    career_id: int = None, gestion_id: int = None, report_type: str = "table", status_filter: str = None
):
    db = SessionLocal()
    try:
        filename = f"report_{uuid.uuid4().hex}.pdf"
        filepath = os.path.join(REPORTS_DIR, filename)

        if report_type in ["seguimiento", "seguimiento-cumplimiento"]:
            build_seguimiento_pdf(filepath, career_id, gestion_id, db)
        elif report_type == "conflict":
            conflicts = conflict_service.find_conflicts(db, career_id, gestion_id)
            build_conflict_pdf(filepath, conflicts, career_id, gestion_id)
        elif report_type in ["agenda-completa", "agenda-academica", "agenda-cientifica", "research-agenda"]:
            career = (
                db.query(Career).filter(Career.id == career_id).first()
                if career_id
                else None
            )
            gestion = (
                db.query(Gestion).filter(Gestion.id == gestion_id).first()
                if gestion_id
                else None
            )
            career_name = career.name if career else "Todas las carreras"
            gestion_name = gestion.name if gestion else "Todas las gestiones"

            activities_sc = []
            if report_type in ["agenda-completa", "agenda-cientifica", "research-agenda"]:
                query_sc = db.query(ScientificActivity)
                if career_id is not None:
                    query_sc = query_sc.filter(ScientificActivity.career_id == career_id)
                if gestion_id is not None:
                    query_sc = query_sc.filter(ScientificActivity.gestion_id == gestion_id)
                if status_filter:
                    query_sc = query_sc.filter(ScientificActivity.status == status_filter)
                activities_sc = query_sc.all()

            activities_ac = []
            if report_type in ["agenda-completa", "agenda-academica"] and (not status_filter or status_filter == 'scheduled'):
                query_ac = db.query(AcademicActivity)
                if career_id is not None:
                    query_ac = query_ac.filter(AcademicActivity.career_id == career_id)
                if gestion_id is not None:
                    query_ac = query_ac.filter(AcademicActivity.gestion_id == gestion_id)
                activities_ac = query_ac.all()
            
            class UnifiedActivity:
                def __init__(self, obj, is_scientific=True):
                    self.title = obj.title
                    self.start_date = obj.start_date
                    self.end_date = obj.end_date
                    self.activity_type = obj.activity_type if is_scientific else obj.category
                    self.is_scientific = is_scientific

            activities = [UnifiedActivity(a, True) for a in activities_sc] + [UnifiedActivity(a, False) for a in activities_ac]
            activities.sort(key=lambda x: x.start_date)

            if report_type == "agenda-academica":
                r_title = "Calendario Académico"
            elif report_type in ["agenda-cientifica", "research-agenda"]:
                r_title = "Calendario de Investigación"
            else:
                r_title = "Calendario Académico y Científico"

            build_research_agenda_pdf(filepath, activities, career_name, gestion_name, r_title)
        else:
            _build_table_report(filepath, career_id, gestion_id, db, status_filter)

        return {"status": "completed", "file_path": filepath, "file_name": filename}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"status": "failed", "error": str(e)}
    finally:
        db.close()


@celery_app.task
def generate_excel_report_task(
    career_id: int = None, gestion_id: int = None, report_type: str = "table", status_filter: str = None
):
    db = SessionLocal()
    try:
        if report_type == "conflict":
            conflicts = conflict_service.find_conflicts(db, career_id, gestion_id)
            return build_conflict_excel(conflicts, career_id, gestion_id)

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        filename = f"report_{uuid.uuid4().hex}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)

        if report_type in ["seguimiento", "seguimiento-cumplimiento"]:
            data = build_seguimiento_data(db, career_id, gestion_id)
            wb = Workbook()
            
            # Sheet 1: Resumen Cumplimiento
            ws_summary = wb.active
            ws_summary.title = "Resumen por Carrera"
            
            headers = ["Carrera / Alcance", "Total Actividades", "Programadas", "En Desarrollo", "Completadas", "Canceladas", "% Cumplimiento"]
            ws_summary.append(headers)
            
            for row in data["careers_summary"]:
                ws_summary.append([
                    row["career_name"],
                    row["total"],
                    row["scheduled"],
                    row["in_progress"],
                    row["completed"],
                    row["cancelled"],
                    f"{row['completion_rate']}%"
                ])
                
            tot = data["totals"]
            ws_summary.append([
                "TOTAL INSTITUCIONAL",
                tot["total"],
                tot["scheduled"],
                tot["in_progress"],
                tot["completed"],
                tot["cancelled"],
                f"{tot['completion_rate']}%"
            ])
            
            # Sheet 2: Detalle Actividades
            ws_detail = wb.create_sheet("Detalle Actividades")
            ws_detail.append(["Título", "Carrera", "Fechas", "Tipo / Categoría", "Estado", "Observaciones / Evidencias"])
            for act in data["detailed_activities"]:
                obs = []
                if act["notes"]:
                    obs.append(f"Notas: {act['notes']}")
                if act["has_evidence"]:
                    obs.append("Evidencia adjunta")
                ws_detail.append([
                    act["title"],
                    act["career_name"],
                    act["dates"],
                    act["type_label"],
                    act["status_label"],
                    " | ".join(obs) if obs else "—"
                ])
                
            wb.save(filepath)
            return {"status": "completed", "file_path": filepath, "file_name": filename}

        ac_query = db.query(AcademicActivity)
        sc_query = db.query(ScientificActivity)

        if career_id is not None:
            ac_query = ac_query.filter(AcademicActivity.career_id == career_id)
            sc_query = sc_query.filter(ScientificActivity.career_id == career_id)
        if gestion_id is not None:
            ac_query = ac_query.filter(AcademicActivity.gestion_id == gestion_id)
            sc_query = sc_query.filter(ScientificActivity.gestion_id == gestion_id)
        if status_filter:
            sc_query = sc_query.filter(ScientificActivity.status == status_filter)

        academic_activities = ac_query.all() if (not status_filter or status_filter == 'scheduled') else []
        scientific_activities = sc_query.all()

        wb = Workbook()

        # Academic sheet
        if academic_activities:
            ws_ac = wb.active
            ws_ac.title = "Académicas"
            ws_ac.append(["Título", "Fecha inicio", "Fecha fin", "Categoría", "Carrera ID"])
            for act in academic_activities:
                ws_ac.append([
                    act.title,
                    act.start_date.isoformat() if act.start_date else "",
                    act.end_date.isoformat() if act.end_date else "",
                    act.category or "",
                    act.career_id,
                ])
            ws_sc = wb.create_sheet("Científicas")
        else:
            ws_sc = wb.active
            ws_sc.title = "Científicas"

        ws_sc.append(["Título", "Fecha inicio", "Fecha fin", "Tipo", "Estado", "Responsable", "Evidencias / Respaldos", "Observaciones / Justificación", "Carrera ID"])
        for act in scientific_activities:
            has_ev = (len(act.evidences) > 0 or bool(act.evidence_url)) if hasattr(act, 'evidences') else False
            ws_sc.append([
                act.title,
                act.start_date.isoformat() if act.start_date else "",
                act.end_date.isoformat() if act.end_date else "",
                get_activity_label(act.activity_type),
                get_status_label(act.status),
                act.responsible_name or "",
                f"✓ Con evidencia ({len(act.evidences)} archivos)" if (hasattr(act, 'evidences') and len(act.evidences) > 0) else ("✓ Con evidencia" if has_ev else "Sin evidencia"),
                act.notes or "—",
                act.career_id,
            ])

        wb.save(filepath)

        return {"status": "completed", "file_path": filepath, "file_name": filename}
    except Exception as e:
        return {"status": "failed", "error": str(e)}
    finally:
        db.close()
