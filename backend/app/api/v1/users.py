import math
import re
from io import BytesIO
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, status, Response
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.models.models import User, Career
from app.schemas.schemas import (
    UserResponse,
    UserCreate,
    UserUpdate,
    UserAdminUpdate,
    PaginatedUserResponse,
    UserImportReport,
    UserImportRowError,
    RoleEnum,
)
from app.api.deps import get_current_active_user, require_admin_role
from app.core.security import get_password_hash
from app.workers.notification_worker import send_telegram_message

router = APIRouter()

ROLE_MAP = {
    "docente": RoleEnum.teacher,
    "teacher": RoleEnum.teacher,
    "coordinador": RoleEnum.coordinator,
    "coordinator": RoleEnum.coordinator,
    "investigador": RoleEnum.research,
    "research": RoleEnum.research,
    "administrador": RoleEnum.admin,
    "admin": RoleEnum.admin,
    "super admin": RoleEnum.super_admin,
    "super_admin": RoleEnum.super_admin,
    "superadmin": RoleEnum.super_admin,
    "lectura": RoleEnum.read_only,
    "solo lectura": RoleEnum.read_only,
    "read_only": RoleEnum.read_only,
    "vicerrectorado": RoleEnum.vicerrectorado,
    "director de investigación": RoleEnum.director_investigacion,
    "director de investigacion": RoleEnum.director_investigacion,
    "director_investigacion": RoleEnum.director_investigacion,
    "jefe de investigación": RoleEnum.jefe_investigacion,
    "jefe de investigacion": RoleEnum.jefe_investigacion,
    "jefe_investigacion": RoleEnum.jefe_investigacion,
}

for r in RoleEnum:
    ROLE_MAP[r.value.lower()] = r


def resolve_role(role_val: str) -> RoleEnum:
    if not role_val:
        return RoleEnum.teacher
    val_lower = role_val.strip().lower()
    if val_lower in ROLE_MAP:
        return ROLE_MAP[val_lower]
    for part in val_lower.split("-"):
        p = part.strip()
        if p in ROLE_MAP:
            return ROLE_MAP[p]
    return RoleEnum.teacher


def extract_career_ids(careers_val: str) -> List[int]:
    cids = []
    if not careers_val:
        return cids
    for token in str(careers_val).split(","):
        token = token.strip()
        if not token:
            continue
        match = re.search(r'^\s*(\d+)', token)
        if match:
            cids.append(int(match.group(1)))
    return cids


@router.get("/me", response_model=UserResponse)
def read_user_me(
    current_user: User = Depends(get_current_active_user),
) -> User:
    """
    Get current user.
    """
    return current_user


@router.patch("/me", response_model=UserResponse)
def update_user_me(
    *,
    db: Session = Depends(get_db),
    user_in: UserUpdate,
    current_user: User = Depends(get_current_active_user),
) -> User:
    """
    Update current user's profile (name, phone, telegram chat id, email).
    """
    update_data = user_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(current_user, field, value)

    db.add(current_user)
    db.commit()
    db.refresh(current_user)
    return current_user


@router.post("/me/test-telegram", status_code=status.HTTP_200_OK)
def test_telegram_notification(
    current_user: User = Depends(get_current_active_user),
):
    """Send a test message to the current user's Telegram chat ID."""
    if not current_user.telegram_chat_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No tenés un Telegram Chat ID configurado. Guardalo primero en tu perfil.",
        )

    message = (
        f"Hola {current_user.full_name or current_user.email},\n\n"
        "Este es un mensaje de prueba del Sistema Agenda Científica UNITEPC.\n"
        "Si llegaste a leer esto, tu bot de Telegram está funcionando correctamente."
    )

    sent = send_telegram_message(current_user.telegram_chat_id, message)
    if not sent:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo enviar el mensaje de prueba. Verificá que el TELEGRAM_BOT_TOKEN esté configurado.",
        )

    return {"message": "Mensaje de prueba enviado correctamente por Telegram"}


@router.get("/", response_model=PaginatedUserResponse)
def list_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    role: Optional[RoleEnum] = Query(None),
    career_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_role),
):
    """
    List users with pagination, search, role, and career filters.
    """
    query = db.query(User)

    if search:
        search_fmt = f"%{search.strip()}%"
        query = query.filter(
            or_(
                User.email.ilike(search_fmt),
                User.full_name.ilike(search_fmt),
            )
        )

    if role:
        query = query.filter(User.role == role)

    if career_id:
        query = query.filter(User.careers.any(Career.id == career_id))

    total = query.count()
    pages = math.ceil(total / page_size) if total > 0 else 1
    items = query.order_by(User.id.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return PaginatedUserResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=pages,
    )


@router.get("/excel-template")
def get_user_excel_template(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_role),
):
    """
    Generates and returns an Excel (.xlsx) template for bulk user importing.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # Title / headers
    ws.append(["Email *", "Nombre Completo", "Rol", "Teléfono", "Telegram Chat ID", "Contraseña", "IDs Carreras (separadas por coma)"])

    careers = db.query(Career).order_by(Career.id).all()
    if careers:
        career_options = [f"{c.id} - {c.name}" for c in careers]
    else:
        career_options = ["1 - Carrera General"]

    sample_career = career_options[0]
    ws.append(["ejemplo.docente@unitepc.edu.bo", "Juan Pérez", "Docente", "+59170000000", "12345678", "Unitepc2026!", sample_career])
    ws.append(["coordinador.med@unitepc.edu.bo", "María Lopez", "Coordinador", "+59171111111", "", "Unitepc2026!", sample_career])

    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    widths = [35, 25, 18, 16, 20, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Catalog sheet creation
    ws_cat = wb.create_sheet(title="Catalogos")
    roles = [
        "Docente",
        "Coordinador",
        "Investigador",
        "Administrador",
        "Super Admin",
        "Lectura",
        "Vicerrectorado",
        "Director de Investigación",
        "Jefe de Investigación",
    ]

    ws_cat.append(["Roles", "Carreras"])
    max_rows = max(len(roles), len(career_options))
    for i in range(max_rows):
        r_val = roles[i] if i < len(roles) else None
        c_val = career_options[i] if i < len(career_options) else None
        ws_cat.append([r_val, c_val])

    # DataValidation setup on Usuarios sheet
    dv_role = DataValidation(type="list", formula1="=Catalogos!$A$2:$A$10", allow_blank=True)
    ws.add_data_validation(dv_role)
    dv_role.add("C2:C500")

    N = 1 + len(career_options)
    dv_career = DataValidation(type="list", formula1=f"=Catalogos!$B$2:$B${N}", allow_blank=True)
    dv_career.showErrorMessage = False
    ws.add_data_validation(dv_career)
    dv_career.add("G2:G500")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": "attachment; filename=plantilla_importacion_usuarios.xlsx"
    }
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/import-excel", response_model=UserImportReport)
def import_users_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_role),
):
    """
    Bulk import users from an uploaded Excel (.xlsx) file.
    """
    import openpyxl

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Formato de archivo inválido. Debe subirse un archivo Excel (.xlsx)."
        )

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {str(e)}")

    ws = wb["Usuarios"] if "Usuarios" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return UserImportReport(total_rows=0, success_count=0, error_count=0, row_errors=[])

    # Determine header row index
    header_idx = 0
    for idx, r in enumerate(rows):
        row_vals = [str(cell).lower().strip() for cell in r if cell is not None]
        if any("email" in val for val in row_vals):
            header_idx = idx
            break

    data_rows = rows[header_idx + 1:]
    total_rows = len(data_rows)
    success_count = 0
    row_errors = []

    existing_emails = set(e[0].lower() for e in db.query(User.email).all())
    processed_emails_in_batch = set()

    for idx, r in enumerate(data_rows, start=header_idx + 2):
        if not r or all(cell is None or str(cell).strip() == "" for cell in r):
            total_rows -= 1
            continue

        email_val = str(r[0]).strip() if len(r) > 0 and r[0] is not None else ""
        name_val = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        role_val = str(r[2]).strip() if len(r) > 2 and r[2] is not None else "teacher"
        phone_val = str(r[3]).strip() if len(r) > 3 and r[3] is not None else None
        telegram_val = str(r[4]).strip() if len(r) > 4 and r[4] is not None else None
        password_val = str(r[5]).strip() if len(r) > 5 and r[5] is not None else "Unitepc2026!"
        careers_val = str(r[6]).strip() if len(r) > 6 and r[6] is not None else ""

        if not email_val or "@" not in email_val:
            row_errors.append(UserImportRowError(row=idx, email=email_val or None, error="Email inválido o ausente"))
            continue

        email_lower = email_val.lower()

        if email_lower in existing_emails or email_lower in processed_emails_in_batch:
            row_errors.append(UserImportRowError(row=idx, email=email_val, error="El email ya se encuentra registrado"))
            continue

        role_enum_val = resolve_role(role_val)

        if role_enum_val == RoleEnum.super_admin and current_user.role != RoleEnum.super_admin:
            row_errors.append(UserImportRowError(row=idx, email=email_val, error="No tiene permisos para importar usuarios super_admin"))
            continue
        if role_enum_val == RoleEnum.admin and current_user.role not in (RoleEnum.super_admin, RoleEnum.admin):
            row_errors.append(UserImportRowError(row=idx, email=email_val, error="No tiene permisos para importar usuarios admin"))
            continue

        career_objects = []
        if careers_val:
            valid_cids = extract_career_ids(careers_val)
            if valid_cids:
                career_objects = db.query(Career).filter(Career.id.in_(valid_cids)).all()

        try:
            new_user = User(
                email=email_val,
                full_name=name_val or None,
                role=role_enum_val,
                phone_number=phone_val or None,
                telegram_chat_id=telegram_val or None,
                hashed_password=get_password_hash(password_val if password_val else "Unitepc2026!"),
                is_active=True,
            )
            if career_objects:
                new_user.careers.extend(career_objects)

            db.add(new_user)
            db.flush()
            processed_emails_in_batch.add(email_lower)
            success_count += 1
        except Exception as ex:
            db.rollback()
            row_errors.append(UserImportRowError(row=idx, email=email_val, error=f"Error insertando usuario: {str(ex)}"))

    db.commit()

    return UserImportReport(
        total_rows=total_rows,
        success_count=success_count,
        error_count=len(row_errors),
        row_errors=row_errors,
    )


@router.post("/", response_model=UserResponse)
def create_user(
    *,
    db: Session = Depends(get_db),
    user_in: UserCreate,
    current_user: User = Depends(require_admin_role),
) -> User:
    """
    Create new user.
    """
    if user_in.role == RoleEnum.admin and current_user.role != RoleEnum.super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only super_admin can create admin users"
        )
    if user_in.role == RoleEnum.super_admin and current_user.role != RoleEnum.super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only super_admin can create super_admin users"
        )

    user = db.query(User).filter(User.email == user_in.email).first()
    if user:
        raise HTTPException(
            status_code=400,
            detail="The user with this email already exists in the system.",
        )
        
    user = User(
        email=user_in.email,
        hashed_password=get_password_hash(user_in.password),
        full_name=user_in.full_name,
        role=user_in.role,
        is_active=True,
    )
    
    if user_in.career_ids:
        careers = db.query(Career).filter(Career.id.in_(user_in.career_ids)).all()
        user.careers.extend(careers)

    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@router.put("/{user_id}", response_model=UserResponse)
def update_user_admin(
    user_id: int,
    user_in: UserAdminUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_role),
):
    """
    Update a user's admin details (role, career assignments, active status, name, etc.).
    Includes privilege escalation guard.
    """
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if user.role == RoleEnum.super_admin and current_user.role != RoleEnum.super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only super_admin can modify super_admin users",
        )

    if user_in.role is not None:
        if user_in.role == RoleEnum.super_admin and current_user.role != RoleEnum.super_admin:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only super_admin can assign the super_admin role",
            )
        if user_in.role == RoleEnum.admin and current_user.role not in (RoleEnum.super_admin, RoleEnum.admin):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only super_admin or admin can assign the admin role",
            )

    update_data = user_in.model_dump(exclude_unset=True)

    if "password" in update_data:
        pwd = update_data.pop("password")
        if pwd:
            user.hashed_password = get_password_hash(pwd)

    if "career_ids" in update_data:
        career_ids = update_data.pop("career_ids")
        if career_ids is not None:
            careers = db.query(Career).filter(Career.id.in_(career_ids)).all()
            user.careers = careers

    for field, value in update_data.items():
        if value is not None:
            setattr(user, field, value)

    db.add(user)
    db.commit()
    db.refresh(user)
    return user

