from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, date
import pandas as pd
from pydantic import ValidationError
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db.session import get_db
from app.api.deps import require_read_only_get, check_activity_scope_permission
from app.models.models import AcademicActivity, ScientificActivity, Career, Gestion, ActivityCategory, User
from app.schemas.schemas import ActivityRowValidator, ScientificActivityType

router = APIRouter()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_header_row(ws, headers: list, fill_color: str):
    """Apply header styling to the first row."""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        bottom=Side(border_style="thin", color="CCCCCC"),
        right=Side(border_style="thin", color="CCCCCC"),
    )
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.row_dimensions[1].height = 30


def _add_example_row(ws, values: list):
    """Add a light-gray example row."""
    fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    font = Font(color="64748B", italic=True, size=10)
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Template Download
# ---------------------------------------------------------------------------

@router.get("/template/download")
def download_template(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_read_only_get),
):
    """Return a bilingual blank Excel template with career/gestion/category reference and combobox data validation."""
    wb = openpyxl.Workbook()

    # ---- fetch live data ----
    careers = db.query(Career).order_by(Career.id).all()
    gestiones = db.query(Gestion).order_by(Gestion.id).all()
    categories = db.query(ActivityCategory).filter(ActivityCategory.is_active.is_(True)).order_by(ActivityCategory.id).all()

    # ---- Sheet 1: Academic Activities ----
    ws_ac = wb.active
    ws_ac.title = "Actividades Académicas"
    academic_headers = [
        "titulo",
        "fecha_inicio",
        "fecha_fin",
        "categoria",
        "carrera",
        "gestion",
        "es_cientifica",
    ]
    _style_header_row(ws_ac, academic_headers, "009E96")
    _add_example_row(ws_ac, [
        "Inicio de clases",
        "11/02/2026",
        "11/02/2026",
        categories[0].name if categories else "GENERAL",
        careers[0].name if careers else "Tec. Superior Prótesis Dental",
        gestiones[0].name if gestiones else "2-2026",
        "NO",
    ])
    ws_ac.cell(row=3, column=1, value="⚠ La fila 2 es ejemplo – bórrala antes de importar")
    ws_ac.cell(row=3, column=1).font = Font(color="EF4444", bold=True, size=9)

    # Format date columns B and C as DD/MM/YYYY
    for r in range(2, 501):
        ws_ac.cell(row=r, column=2).number_format = "DD/MM/YYYY"
        ws_ac.cell(row=r, column=3).number_format = "DD/MM/YYYY"

    # ---- Sheet 2: Scientific Activities ----
    ws_sc = wb.create_sheet(title="Actividades Científicas")
    scientific_headers = [
        "titulo",
        "fecha_inicio",
        "fecha_fin",
        "tipo_actividad",
        "nombre_responsable",
        "carrera",
        "gestion",
        "es_cientifica",
    ]
    _style_header_row(ws_sc, scientific_headers, "6B3392")
    _add_example_row(ws_sc, [
        "Congreso de Investigación",
        "15/03/2026",
        "17/03/2026",
        "CONGRESO",
        "Dr. Juan Pérez",
        careers[0].name if careers else "Tec. Superior Prótesis Dental",
        gestiones[0].name if gestiones else "2-2026",
        "SI",
    ])
    ws_sc.cell(row=3, column=1, value="⚠ La fila 2 es ejemplo – bórrala antes de importar")
    ws_sc.cell(row=3, column=1).font = Font(color="EF4444", bold=True, size=9)

    # Format date columns B and C as DD/MM/YYYY
    for r in range(2, 501):
        ws_sc.cell(row=r, column=2).number_format = "DD/MM/YYYY"
        ws_sc.cell(row=r, column=3).number_format = "DD/MM/YYYY"

    # ---- Sheet 3: Reference (fields + careers + gestiones + categories) ----
    ws_ref = wb.create_sheet(title="Referencia")

    # --- Column widths ---
    ws_ref.column_dimensions["A"].width = 22
    ws_ref.column_dimensions["B"].width = 55
    ws_ref.column_dimensions["D"].width = 10
    ws_ref.column_dimensions["E"].width = 35
    ws_ref.column_dimensions["F"].width = 25
    ws_ref.column_dimensions["H"].width = 10
    ws_ref.column_dimensions["I"].width = 22
    ws_ref.column_dimensions["K"].width = 15
    ws_ref.column_dimensions["L"].width = 25
    ws_ref.column_dimensions["M"].width = 15

    dark_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    green_fill = PatternFill(start_color="009E96", end_color="009E96", fill_type="solid")
    purple_fill = PatternFill(start_color="6B3392", end_color="6B3392", fill_type="solid")
    blue_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    alt_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    # --- Fields section header ---
    ws_ref.cell(row=1, column=1, value="📋 Campos del archivo Excel").font = white_bold
    ws_ref.cell(row=1, column=1).fill = dark_fill
    ws_ref.cell(row=1, column=1).alignment = center
    ws_ref.merge_cells("A1:B1")
    ws_ref.row_dimensions[1].height = 24

    # sub-header
    for col, txt in [(1, "Columna (nombre exacto)"), (2, "Descripción y valores válidos")]:
        c = ws_ref.cell(row=2, column=col, value=txt)
        c.fill = green_fill
        c.font = white_bold
        c.alignment = center

    field_rows = [
        ("titulo", "Nombre de la actividad (texto libre, requerido)"),
        ("fecha_inicio", "Fecha inicio: formato DD/MM/AAAA (ej. 11/02/2026) o YYYY-MM-DD"),
        ("fecha_fin", "Fecha fin: formato DD/MM/AAAA (igual a inicio si dura 1 día)"),
        ("categoria", "Solo académicas: selecciona de la lista desplegable o escribe la categoría"),
        ("tipo_actividad", "Solo científicas: selecciona del desplegable (CONGRESO, WEBINAR, etc.)"),
        ("nombre_responsable", "Solo científicas: nombre del docente o investigador responsable"),
        ("carrera", "Selecciona el nombre de la carrera de la lista desplegable"),
        ("gestion", "Selecciona la gestión de la lista desplegable (ej. 2-2026)"),
        ("es_cientifica", "Selecciona SI para científicas | NO para académicas"),
    ]
    for r_off, (col_a, col_b) in enumerate(field_rows, start=3):
        row = r_off
        ca = ws_ref.cell(row=row, column=1, value=col_a)
        cb = ws_ref.cell(row=row, column=2, value=col_b)
        ca.font = Font(bold=True, color="009E96", size=10)
        cb.font = Font(color="334155", size=10)
        if r_off % 2 == 0:
            ca.fill = alt_fill
            cb.fill = alt_fill

    # --- Careers table (column D-F) ---
    ws_ref.cell(row=1, column=4, value="🎓 Carreras").font = white_bold
    ws_ref.cell(row=1, column=4).fill = green_fill
    ws_ref.cell(row=1, column=4).alignment = center
    ws_ref.merge_cells("D1:F1")

    for col, txt in [(4, "ID"), (5, "Nombre de la Carrera"), (6, "Facultad")]:
        c = ws_ref.cell(row=2, column=col, value=txt)
        c.fill = green_fill
        c.font = white_bold
        c.alignment = center

    for r_off, career in enumerate(careers, start=3):
        row = r_off
        for col, val in [(4, career.id), (5, career.name), (6, career.faculty)]:
            c = ws_ref.cell(row=row, column=col, value=val)
            c.font = Font(color="1E293B", size=10)
            c.alignment = Alignment(horizontal="center" if col == 4 else "left", vertical="center")
            if r_off % 2 == 0:
                c.fill = alt_fill

    # --- Gestiones table (column H-I) ---
    ws_ref.cell(row=1, column=8, value="📅 Gestiones").font = white_bold
    ws_ref.cell(row=1, column=8).fill = purple_fill
    ws_ref.cell(row=1, column=8).alignment = center
    ws_ref.merge_cells("H1:I1")

    for col, txt in [(8, "ID"), (9, "Nombre de la Gestión")]:
        c = ws_ref.cell(row=2, column=col, value=txt)
        c.fill = purple_fill
        c.font = white_bold
        c.alignment = center

    for r_off, gestion in enumerate(gestiones, start=3):
        row = r_off
        for col, val in [(8, gestion.id), (9, gestion.name)]:
            c = ws_ref.cell(row=row, column=col, value=val)
            c.font = Font(color="1E293B", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if r_off % 2 == 0:
                c.fill = alt_fill

    # --- Categories table (column K-M) ---
    ws_ref.cell(row=1, column=11, value="🏷 Categorías Dinámicas").font = white_bold
    ws_ref.cell(row=1, column=11).fill = blue_fill
    ws_ref.cell(row=1, column=11).alignment = center
    ws_ref.merge_cells("K1:M1")

    for col, txt in [(11, "Código"), (12, "Nombre"), (13, "Ámbito")]:
        c = ws_ref.cell(row=2, column=col, value=txt)
        c.fill = blue_fill
        c.font = white_bold
        c.alignment = center

    for r_off, cat in enumerate(categories, start=3):
        row = r_off
        for col, val in [(11, cat.code), (12, cat.name), (13, cat.scope)]:
            c = ws_ref.cell(row=row, column=col, value=val)
            c.font = Font(color="1E293B", size=10)
            c.alignment = Alignment(horizontal="center" if col in (11, 13) else "left", vertical="center")
            if r_off % 2 == 0:
                c.fill = alt_fill

    # ---- Add Data Validations (Comboboxes + Date Validation) ----
    dv_date_ac = DataValidation(type="date", operator="greaterThanOrEqual", formula1="2000-01-01", allow_blank=True)
    dv_date_ac.error = "Formato de fecha inválido. Ingresa DD/MM/AAAA (ej. 11/02/2026)"
    dv_date_ac.errorTitle = "Fecha Inválida"
    ws_ac.add_data_validation(dv_date_ac)
    dv_date_ac.add("B2:C500")

    dv_date_sc = DataValidation(type="date", operator="greaterThanOrEqual", formula1="2000-01-01", allow_blank=True)
    dv_date_sc.error = "Formato de fecha inválido. Ingresa DD/MM/AAAA (ej. 15/03/2026)"
    dv_date_sc.errorTitle = "Fecha Inválida"
    ws_sc.add_data_validation(dv_date_sc)
    dv_date_sc.add("B2:C500")

    dv_bool_ac = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
    ws_ac.add_data_validation(dv_bool_ac)
    dv_bool_ac.add("G2:G500")

    dv_bool_sc = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
    ws_sc.add_data_validation(dv_bool_sc)
    dv_bool_sc.add("H2:H500")

    if careers:
        career_ref = f"'Referencia'!E3:E{2 + len(careers)}"
        dv_car_ac = DataValidation(type="list", formula1=career_ref, allow_blank=True)
        ws_ac.add_data_validation(dv_car_ac)
        dv_car_ac.add("E2:E500")

        dv_car_sc = DataValidation(type="list", formula1=career_ref, allow_blank=True)
        ws_sc.add_data_validation(dv_car_sc)
        dv_car_sc.add("F2:F500")

    if gestiones:
        gestion_ref = f"'Referencia'!I3:I{2 + len(gestiones)}"
        dv_ges_ac = DataValidation(type="list", formula1=gestion_ref, allow_blank=True)
        ws_ac.add_data_validation(dv_ges_ac)
        dv_ges_ac.add("F2:F500")

        dv_ges_sc = DataValidation(type="list", formula1=gestion_ref, allow_blank=True)
        ws_sc.add_data_validation(dv_ges_sc)
        dv_ges_sc.add("G2:G500")

    if categories:
        cat_ref = f"'Referencia'!L3:L{2 + len(categories)}"
        dv_cat_ac = DataValidation(type="list", formula1=cat_ref, allow_blank=True)
        ws_ac.add_data_validation(dv_cat_ac)
        dv_cat_ac.add("D2:D500")

        dv_cat_sc = DataValidation(type="list", formula1=cat_ref, allow_blank=True)
        ws_sc.add_data_validation(dv_cat_sc)
        dv_cat_sc.add("D2:D500")
    else:
        dv_types = DataValidation(type="list", formula1='"CONGRESO,WEBINAR,DEFENSA,FERIA,OLIMPIADA,MASTER_CLASS"', allow_blank=True)
        ws_sc.add_data_validation(dv_types)
        dv_types.add("D2:D500")

    # ---- Stream ----
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_actividades.xlsx"},
    )


# ---------------------------------------------------------------------------
# Upload Excel
# ---------------------------------------------------------------------------

def parse_date_val(val):
    if val is None or pd.isna(val):
        return None
    if isinstance(val, (datetime, date)):
        if isinstance(val, datetime):
            return val.date()
        return val
    val_str = str(val).strip()
    try:
        dt = pd.to_datetime(val_str, dayfirst=True)
        return dt.date()
    except Exception:
        return val_str


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_read_only_get),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Invalid file format. Only Excel files are supported.")
    
    try:
        contents = await file.read()
        excel_sheets = pd.read_excel(BytesIO(contents), sheet_name=None)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")
        
    raw_records = []
    for sheet_name, df in excel_sheets.items():
        if sheet_name.strip().lower() in ["referencia", "reference"]:
            continue
        sheet_recs = df.to_dict('records')
        for rec in sheet_recs:
            # Skip warning / example hint rows
            title_val = str(rec.get("titulo", rec.get("title", "")) or "")
            if "bórrala antes de importar" in title_val.lower() or "⚠" in title_val:
                continue
            
            # Infer is_scientific from sheet name if not specified
            if "es_cientifica" not in rec and "is_scientific" not in rec:
                if "científica" in sheet_name.lower() or "cientifica" in sheet_name.lower():
                    rec["es_cientifica"] = True
                elif "académica" in sheet_name.lower() or "academica" in sheet_name.lower():
                    rec["es_cientifica"] = False
            raw_records.append(rec)

    records = raw_records

    # Accept Spanish column names (from the downloadable template)
    COLUMN_MAP = {
        "titulo": "title",
        "fecha_inicio": "start_date",
        "fecha_fin": "end_date",
        "categoria": "category",
        "tipo_actividad": "activity_type",
        "nombre_responsable": "responsible_name",
        "id_carrera": "career_id",
        "carrera": "career_id",
        "id_gestion": "gestion_id",
        "gestion": "gestion_id",
        "es_cientifica": "is_scientific",
    }

    # Build Carrera lookup dictionary
    careers_all = db.query(Career).all()
    career_map = {}
    for c in careers_all:
        career_map[str(c.id)] = c.id
        career_map[c.name.strip().upper()] = c.id
        career_map[f"{c.id} - {c.name}".strip().upper()] = c.id
        career_map[f"{c.name} ({c.id})".strip().upper()] = c.id

    # Build Gestion lookup dictionary
    gestiones_all = db.query(Gestion).all()
    gestion_map = {}
    for g in gestiones_all:
        gestion_map[str(g.id)] = g.id
        gestion_map[g.name.strip().upper()] = g.id
        gestion_map[f"{g.id} - {g.name}".strip().upper()] = g.id
        gestion_map[f"{g.name} ({g.id})".strip().upper()] = g.id

    renamed_records = []
    for record in records:
        renamed = {COLUMN_MAP.get(str(k).strip().lower(), k): v for k, v in record.items()}

        # Normalize start_date & end_date (supports DD/MM/YYYY and YYYY-MM-DD)
        if "start_date" in renamed and pd.notnull(renamed["start_date"]):
            renamed["start_date"] = parse_date_val(renamed["start_date"])
        if "end_date" in renamed and pd.notnull(renamed["end_date"]):
            renamed["end_date"] = parse_date_val(renamed["end_date"])

        # Normalize career_id / carrera
        if "career_id" in renamed and pd.notnull(renamed["career_id"]):
            raw = str(renamed["career_id"]).strip().upper()
            if raw.endswith(".0"):
                raw = raw[:-2]
            if raw in career_map:
                renamed["career_id"] = career_map[raw]

        # Normalize gestion_id / gestion
        if "gestion_id" in renamed and pd.notnull(renamed["gestion_id"]):
            raw = str(renamed["gestion_id"]).strip().upper()
            if raw.endswith(".0"):
                raw = raw[:-2]
            if raw in gestion_map:
                renamed["gestion_id"] = gestion_map[raw]

        # Normalize is_scientific
        if "is_scientific" in renamed and pd.notnull(renamed["is_scientific"]):
            raw_b = renamed["is_scientific"]
            if isinstance(raw_b, bool):
                renamed["is_scientific"] = raw_b
            else:
                val = str(raw_b).strip().upper()
                renamed["is_scientific"] = val in ["TRUE", "SI", "SÍ", "1", "YES"]

        renamed_records.append(renamed)
    records = renamed_records

    # Dynamic Category lookup map
    categories = db.query(ActivityCategory).filter(ActivityCategory.is_active.is_(True)).all()
    cat_lookup = {}
    for cat in categories:
        cat_lookup[cat.code.upper().strip()] = cat
        cat_lookup[cat.name.upper().strip()] = cat

    valid_enum_types = {t.value for t in ScientificActivityType}

    valid_academic = []
    valid_scientific = []
    errors = []
    
    for i, record in enumerate(records):
        clean_record = {k: v for k, v in record.items() if pd.notnull(v)}
        try:
            validated = ActivityRowValidator(**clean_record)
            
            if validated.is_scientific:
                raw_type = (validated.activity_type.value if hasattr(validated.activity_type, 'value')
                            else str(validated.activity_type or ''))
                if not raw_type and not validated.category:
                    errors.append({"row": i + 2, "error": "activity_type or category is required for scientific activities"})
                    continue

                # Match dynamic category by type or category string
                lookup_key = (raw_type or validated.category or '').strip().upper()
                matched_cat = cat_lookup.get(lookup_key)
                category_id = matched_cat.id if matched_cat else None

                # Determine enum activity_type fallback for DB constraint
                type_lower = raw_type.lower().strip()
                if type_lower in valid_enum_types:
                    act_type_val = type_lower
                elif matched_cat and matched_cat.code.lower() in valid_enum_types:
                    act_type_val = matched_cat.code.lower()
                else:
                    act_type_val = ScientificActivityType.congreso.value

                valid_scientific.append({
                    "title": validated.title,
                    "start_date": validated.start_date,
                    "end_date": validated.end_date,
                    "activity_type": act_type_val,
                    "category_id": category_id,
                    "responsible_name": validated.responsible_name or "Unknown",
                    "career_id": validated.career_id,
                    "gestion_id": validated.gestion_id,
                    "status": "scheduled",
                })
            else:
                if not validated.category:
                    errors.append({"row": i + 2, "error": "category is required for academic activities"})
                    continue

                lookup_key = validated.category.strip().upper()
                matched_cat = cat_lookup.get(lookup_key)
                if matched_cat:
                    category_id = matched_cat.id
                    category_code = matched_cat.code
                else:
                    category_id = None
                    category_code = validated.category

                valid_academic.append({
                    "title": validated.title,
                    "start_date": validated.start_date,
                    "end_date": validated.end_date,
                    "category": category_code,
                    "category_id": category_id,
                    "career_id": validated.career_id,
                    "gestion_id": validated.gestion_id,
                })
        except ValidationError as e:
            errors.append({"row": i + 2, "error": str(e)})
    
    # Enforce scope-aware permissions before inserting
    for row in valid_academic:
        check_activity_scope_permission(current_user, row.get("career_id"))
    for row in valid_scientific:
        check_activity_scope_permission(current_user, row.get("career_id"))

    # Fetch existing activities to detect duplicates and date overlaps
    existing_academics = db.query(AcademicActivity).all()
    existing_scientifics = db.query(ScientificActivity).filter(
        ScientificActivity.status != ScientificActivityStatus.cancelled
    ).all()

    existing_keys = set()
    for a in existing_academics:
        existing_keys.add((a.title.strip().upper(), a.start_date, a.end_date, a.career_id, a.gestion_id))
    for s in existing_scientifics:
        existing_keys.add((s.title.strip().upper(), s.start_date, s.end_date, s.career_id, s.gestion_id))

    def overlaps(start_a, end_a, start_b, end_b):
        return start_a <= end_b and start_b <= end_a

    duplicate_count = 0
    conflicts_detected = []
    filtered_academic = []
    filtered_scientific = []

    # Process Academic rows for duplicates & conflicts
    for row in valid_academic:
        key = (row["title"].strip().upper(), row["start_date"], row["end_date"], row["career_id"], row["gestion_id"])
        if key in existing_keys:
            duplicate_count += 1
            continue
        existing_keys.add(key)
        filtered_academic.append(row)

        c_id = row["career_id"]
        g_id = row["gestion_id"]
        if c_id and g_id:
            for ex in existing_academics:
                if ex.career_id == c_id and ex.gestion_id == g_id and overlaps(row["start_date"], row["end_date"], ex.start_date, ex.end_date):
                    conflicts_detected.append({
                        "activity_title": row["title"],
                        "conflicting_title": ex.title,
                        "dates": f"{ex.start_date} a {ex.end_date}",
                        "career_id": c_id,
                    })
            for ex in existing_scientifics:
                if ex.career_id == c_id and ex.gestion_id == g_id and overlaps(row["start_date"], row["end_date"], ex.start_date, ex.end_date):
                    conflicts_detected.append({
                        "activity_title": row["title"],
                        "conflicting_title": ex.title,
                        "dates": f"{ex.start_date} a {ex.end_date}",
                        "career_id": c_id,
                    })

    # Process Scientific rows for duplicates & conflicts
    for row in valid_scientific:
        key = (row["title"].strip().upper(), row["start_date"], row["end_date"], row["career_id"], row["gestion_id"])
        if key in existing_keys:
            duplicate_count += 1
            continue
        existing_keys.add(key)
        filtered_scientific.append(row)

        c_id = row["career_id"]
        g_id = row["gestion_id"]
        if c_id and g_id:
            for ex in existing_academics:
                if ex.career_id == c_id and ex.gestion_id == g_id and overlaps(row["start_date"], row["end_date"], ex.start_date, ex.end_date):
                    conflicts_detected.append({
                        "activity_title": row["title"],
                        "conflicting_title": ex.title,
                        "dates": f"{ex.start_date} a {ex.end_date}",
                        "career_id": c_id,
                    })
            for ex in existing_scientifics:
                if ex.career_id == c_id and ex.gestion_id == g_id and overlaps(row["start_date"], row["end_date"], ex.start_date, ex.end_date):
                    conflicts_detected.append({
                        "activity_title": row["title"],
                        "conflicting_title": ex.title,
                        "dates": f"{ex.start_date} a {ex.end_date}",
                        "career_id": c_id,
                    })

    inserted_count = 0
    try:
        if filtered_academic:
            db.bulk_insert_mappings(AcademicActivity, filtered_academic)
            inserted_count += len(filtered_academic)
        if filtered_scientific:
            db.bulk_insert_mappings(ScientificActivity, filtered_scientific)
            inserted_count += len(filtered_scientific)
        
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database insertion failed: {str(e)}")
    
    return {
        "inserted_count": inserted_count,
        "duplicate_count": duplicate_count,
        "errors": errors,
        "conflicts": conflicts_detected,
    }
