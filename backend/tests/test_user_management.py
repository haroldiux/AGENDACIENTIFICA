import io
import openpyxl
import pytest
from app.main import app
from app.models.models import RoleEnum, User, Career
from app.api.deps import get_current_active_user, require_admin_role


def _make_admin_user(role: RoleEnum = RoleEnum.super_admin) -> User:
    user = User(
        id=99,
        email="admin@example.com",
        hashed_password="hashed_pw",
        full_name="Admin User",
        role=role,
        is_active=True,
    )
    return user


def test_list_users_paginated_and_filtered(client, db_session):
    # Setup test users
    u1 = User(email="teacher1@unitepc.edu.bo", hashed_password="pw", full_name="Teacher One", role=RoleEnum.teacher, is_active=True)
    u2 = User(email="coordinator1@unitepc.edu.bo", hashed_password="pw", full_name="Coord One", role=RoleEnum.coordinator, is_active=True)
    db_session.add_all([u1, u2])
    db_session.commit()

    admin = _make_admin_user(RoleEnum.admin)
    app.dependency_overrides[get_current_active_user] = lambda: admin
    app.dependency_overrides[require_admin_role] = lambda: admin

    try:
        # List all
        resp = client.get("/api/v1/users/?page=1&page_size=10")
        assert resp.status_code == 200
        data = resp.json()
        assert "items" in data
        assert data["total"] >= 2

        # Search filter
        resp_search = client.get("/api/v1/users/?search=Teacher")
        assert resp_search.status_code == 200
        items = resp_search.json()["items"]
        assert any(u["email"] == "teacher1@unitepc.edu.bo" for u in items)

        # Role filter
        resp_role = client.get("/api/v1/users/?role=coordinator")
        assert resp_role.status_code == 200
        items_role = resp_role.json()["items"]
        assert all(u["role"] == "coordinator" for u in items_role)
    finally:
        app.dependency_overrides.clear()


def test_update_user_admin_and_privilege_guard(client, db_session):
    u = User(email="target@unitepc.edu.bo", hashed_password="pw", full_name="Target User", role=RoleEnum.teacher, is_active=True)
    db_session.add(u)
    db_session.commit()
    db_session.refresh(u)

    # 1. Non-super_admin trying to promote to super_admin should fail (403)
    admin = _make_admin_user(RoleEnum.admin)
    app.dependency_overrides[get_current_active_user] = lambda: admin
    app.dependency_overrides[require_admin_role] = lambda: admin

    try:
        resp = client.put(f"/api/v1/users/{u.id}", json={"role": "super_admin"})
        assert resp.status_code == 403

        # 2. Super_admin updating role and active status
        super_admin = _make_admin_user(RoleEnum.super_admin)
        app.dependency_overrides[require_admin_role] = lambda: super_admin

        resp_ok = client.put(f"/api/v1/users/{u.id}", json={"role": "coordinator", "full_name": "Updated Target", "is_active": False})
        assert resp_ok.status_code == 200
        data = resp_ok.json()
        assert data["role"] == "coordinator"
        assert data["full_name"] == "Updated Target"
        assert data["is_active"] is False
    finally:
        app.dependency_overrides.clear()


def test_get_user_excel_template(client, db_session):
    c1 = Career(name="Sistemas", faculty="Tecnología")
    db_session.add(c1)
    db_session.commit()
    db_session.refresh(c1)

    admin = _make_admin_user(RoleEnum.admin)
    app.dependency_overrides[require_admin_role] = lambda: admin

    try:
        resp = client.get("/api/v1/users/excel-template")
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        # Verify it's a valid openpyxl workbook with Catalogos
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert "Usuarios" in wb.sheetnames
        assert "Catalogos" in wb.sheetnames

        ws_users = wb["Usuarios"]
        assert ws_users.cell(row=1, column=1).value == "Email *"

        ws_cat = wb["Catalogos"]
        assert ws_cat.cell(row=1, column=1).value == "Roles"
        assert ws_cat.cell(row=1, column=2).value == "Carreras"
        assert ws_cat.cell(row=2, column=1).value == "Docente"
        assert ws_cat.cell(row=2, column=2).value == f"{c1.id} - Sistemas"

        # Verify DataValidation on Usuarios worksheet
        dvs = ws_users.data_validations.dataValidation
        assert len(dvs) == 2
        
        role_dv = next((dv for dv in dvs if "C2:C500" in str(dv.sqref)), None)
        career_dv = next((dv for dv in dvs if "G2:G500" in str(dv.sqref)), None)

        assert role_dv is not None
        assert role_dv.formula1 == "=Catalogos!$A$2:$A$10"

        assert career_dv is not None
        assert career_dv.formula1 == "=Catalogos!$B$2:$B$2"
        assert career_dv.showErrorMessage is False
    finally:
        app.dependency_overrides.clear()


def test_get_user_excel_template_zero_careers_fallback(client, db_session):
    db_session.query(Career).delete()
    db_session.commit()

    admin = _make_admin_user(RoleEnum.admin)
    app.dependency_overrides[require_admin_role] = lambda: admin

    try:
        resp = client.get("/api/v1/users/excel-template")
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws_cat = wb["Catalogos"]
        assert ws_cat.cell(row=2, column=2).value == "1 - Carrera General"
    finally:
        app.dependency_overrides.clear()


def test_import_users_excel(client, db_session):
    # Create sample careers
    c1 = Career(name="Medicina Import", faculty="Salud")
    c2 = Career(name="Enfermería Import", faculty="Salud")
    db_session.add_all([c1, c2])
    db_session.commit()
    db_session.refresh(c1)
    db_session.refresh(c2)

    # Create Excel workbook in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(["Email *", "Nombre Completo", "Rol", "Teléfono", "Telegram Chat ID", "Contraseña", "IDs Carreras (separadas por coma)"])
    ws.append(["import.valid@unitepc.edu.bo", "Valid Import User", "Docente", "+59170001111", "998877", "Pass123!", f"{c1.id} - Medicina Import"])
    ws.append(["invalid-email-no-at", "Bad User", "Docente", "", "", "", ""])
    ws.append(["import.valid2@unitepc.edu.bo", "Valid User 2", "Coordinador", "", "", "Pass123!", f"{c1.id} - Medicina Import, {c2.id} - Enfermería Import"])
    ws.append(["import.valid3@unitepc.edu.bo", "Valid User 3", "teacher - Docente", "", "", "Pass123!", ""])

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    admin = _make_admin_user(RoleEnum.super_admin)
    app.dependency_overrides[require_admin_role] = lambda: admin

    try:
        resp = client.post(
            "/api/v1/users/import-excel",
            files={"file": ("import.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert resp.status_code == 200
        report = resp.json()
        assert report["total_rows"] == 4
        assert report["success_count"] == 3
        assert report["error_count"] == 1
        assert len(report["row_errors"]) == 1
        assert report["row_errors"][0]["row"] == 3

        # Verify DB records
        created_user = db_session.query(User).filter(User.email == "import.valid@unitepc.edu.bo").first()
        assert created_user is not None
        assert created_user.full_name == "Valid Import User"
        assert created_user.role == RoleEnum.teacher
        assert len(created_user.careers) == 1
        assert created_user.careers[0].id == c1.id

        created_user2 = db_session.query(User).filter(User.email == "import.valid2@unitepc.edu.bo").first()
        assert created_user2 is not None
        assert created_user2.role == RoleEnum.coordinator
        assert len(created_user2.careers) == 2
        career_ids = {c.id for c in created_user2.careers}
        assert career_ids == {c1.id, c2.id}

        created_user3 = db_session.query(User).filter(User.email == "import.valid3@unitepc.edu.bo").first()
        assert created_user3 is not None
        assert created_user3.role == RoleEnum.teacher
    finally:
        app.dependency_overrides.clear()
